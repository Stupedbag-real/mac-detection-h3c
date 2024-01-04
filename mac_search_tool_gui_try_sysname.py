import customtkinter as tk
# from customtkinter import filedialog
from customtkinter import *
#import tkinter as ttk
import paramiko
import openpyxl
import sys
import time
import re
#import getpass
import io
import logging
import threading
import random
import datetime
import shutil
import ctypes
import subprocess

#logging.getLogger("paramiko").setLevel(logging.DEBUG)
# Create the "mac_debug" folder if it doesn't exist
# Create the "mac_debug" folder if it doesn't exist
if not os.path.exists("mac_debug"):
    os.mkdir("mac_debug")

# Configure the logging for paramiko and other modules
logging.basicConfig(filename="mac_debug/debug.log", level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Get the logger for the current module
logger = logging.getLogger(__name__)


#Create GUI toplevel window
class ToplevelWindow(tk.CTkToplevel):
    def __init__(self, delete_callback, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.geometry("400x300")
        self.title("")
        self.label = tk.CTkLabel(self, text="Delete Debug Files?", font=tk.CTkFont(size=20,weight="bold"))
        self.label.place(x = 100, y = 0 )
        self.tpl_yes_button = tk.CTkButton(self, text="Yes", command=self.delete_files, corner_radius=100)
        self.tpl_yes_button.place(x = 50, y = 50)

        self.tpl_no_button = tk.CTkButton(self, text="No", command=self.close, corner_radius=100)
        self.tpl_no_button.place(x = 200, y = 50)

        self.delete_callback = delete_callback

    def delete_files(self):
        self.delete_callback()
        self.destroy()
    def close(self):
        self.destroy()

# Create GUI
class ConsoleOutput(io.StringIO):
    def __init__(self, textbox):
        self.textbox = textbox
        io.StringIO.__init__(self)

    def write(self, message):
        self.textbox.configure(state='normal')
        self.textbox.insert('end', message)
        self.textbox.see('end')
        self.textbox.configure(state='disabled')

    def flush(self):
        pass


class Application(tk.CTk):
    def __init__(self):
        super().__init__()
        self.title("MAC address search tool")
        self.geometry(f"{1100}x{580}")
        self.stop_threads = False
        # self.minsize(300, 200)
        #self.grid_rowconfigure(0, weight=1)
        #self.grid_columnconfigure(0, weight=1)
        #self.grid_rowconfigure(1, weight=1)
        self.create_widgets()
        #self.set_appearance_mode("system")
        #self.grid_columnconfigure(1, weight=1)
        #self.grid_columnconfigure((2, 3), weight=0)
        #self.grid_rowconfigure((0, 1, 2), weight=1)
        #self.set_widget_scaling()
        #self.bind("<Configure>", self.on_window_resize)  # Bind resize event

        # Make the application DPI aware
        if hasattr(ctypes.windll.shcore, "SetProcessDpiAwareness"):
            ctypes.windll.shcore.SetProcessDpiAwareness(2)
        # Set custom scaling factors for widgets and window dimensions
        tk.set_widget_scaling(1.2)  # Adjust the value as needed
        tk.set_window_scaling(1.2)  # Adjust the value as needed

    # def on_resize(self, event):
    # get the current window size
    # width = self.winfo_width()
    # height = self.winfo_height()

    # check if the size has actually changed
    # if width != self.prev_width or height != self.prev_height:
    # self.prev_width = width
    # self.prev_height = height

    # calculate the new button size
    # button_width = int(width * 0.8)
    # button_height = int(height * 0.1)

    # set the new button size
    # self.username_label.configure(width=button_width, height=button_height)
    # self.username_entry.configure(width=button_width, height=button_height)
    # self.password_label.configure(width=button_width, height=button_height)
    # self.password_entry.configure(width=button_width, height=button_height)
    # self.port_label.configure(width=button_width, height=button_height)
    # self.port_entry.configure(width=button_width, height=button_height)
    # self.mac_address_label.configure(width=button_width, height=button_height)
    # self.mac_address_entry.configure(width=button_width, height=button_height)
    # self.file_path_label.configure(width=button_width, height=button_height)
    # self.browse_button.configure(width=button_width, height=button_height)
    # self.search_button.configure(width=button_width, height=button_height)
    # self.quit_button.configure(width=button_width, height=button_height)

    def create_widgets(self):
        #self.button_1 = tk.CTkButton(self, text="open toplevel", command=self.ask_delete_debug_files)
        #self.button_1.place(x = 450, y = 0 )
        self.toplevel_window = None
        self.var=tk.IntVar()
        #self.device_var=tk.IntVar()
        #self.device_type_radio=tk.CTkRadioButton(self, text="Comware", variable=self.device_var, value=1, command=self.check_comware_version)
        #self.device_type_radio.place(x = 10, y = 10 )
        self.username_label = tk.CTkLabel(self, text="Enter username:", )
        #self.username_label.pack(row=1, column=1, padx=5, pady=5)
        self.username_label.place(x = 850, y = 0 )
        #self.username_label.grid_propagate(False)
        self.username_entry = tk.CTkEntry(self, corner_radius=100)
        #self.username_entry.pack(padx=20, pady=20)
        self.username_entry.place(x = 950, y = 0)
        #self.username_entry.grid_propagate(False)
        self.password_label = tk.CTkLabel(self, text="Enter password:")
        #self.password_label.pack(padx=5, pady=5)
        self.password_label.place(x = 850, y = 50)
        #self.password_label.grid_propagate(False)
        self.password_entry = tk.CTkEntry(self, show="*", corner_radius=100)
        #self.password_entry.pack(padx=20, pady=20)
        self.password_entry.place(y = 50, x = 950)
        #self.password_entry.grid_propagate(False)
        self.port_label = tk.CTkLabel(self, text="Enter physical port to search on:")
        #self.port_label.pack(padx=5, pady=5)
        self.port_label.place(y = 100, x= 760)
        #self.port_label.grid_propagate(False)
        self.port_entry = tk.CTkEntry(self, corner_radius=100)
        #self.port_entry.pack(padx=20, pady=20)
        self.port_entry.place(y = 100, x= 950)
        #self.port_entry.grid_propagate(False)
        self.mac_address_label = tk.CTkLabel(self, text="Enter MAC address to search for:")
        self.mac_address_label.place(y = 150, x = 590)
        #self.mac_address_label.grid_propagate(False)
        # self.mac_address_label.pack(padx=5, pady=5)
        self.mac_address_entry_radio = tk.CTkRadioButton(self, text="Single", variable=self.var, value=1, command=self.toggle_entry_or_browse)
        self.mac_address_browse_radio = tk.CTkRadioButton(self, text="Multiple", variable=self.var, value=2, command=self.toggle_entry_or_browse)
        self.mac_address_entry = tk.CTkEntry(self, corner_radius=100)
        #self.mac_address_entry.pack(padx=20, pady=20)
        self.mac_address_entry_radio.place(y = 150, x = 790)
        self.mac_address_browse_radio.place(y = 150, x = 860)
        #self.mac_address_entry_radio.grid_propagate(False)
        #self.mac_address_browse_radio.grid_propagate(False)
        #self.mac_address_browse_entry.pack(padx=20, pady=20)
        self.mac_address_entry_radio.select()
        self.mac_address_entry.place(y = 150, x = 950)
        #self.mac_address_entry.grid_propagate(False)
        self.mac_address_browse_entry = tk.CTkEntry(self, corner_radius=100)
        self.mac_address_browse_button = tk.CTkButton(self, text="Browse", command=self.browse_file_mac)
        #self.mac_address_browse_button.place(y = 250, x = 560)
        self.file_path_label = tk.CTkLabel(self, text="Path to file with device IP addresses:")
        self.file_path_label.place(y = 200, x = 580)
        #self.file_path_label.grid_propagate(False)
        self.browse_button = tk.CTkButton(self, text="Browse", command=self.browse_device_file)
        self.browse_button.place(y = 200, x = 800)
        #self.browse_button.grid_propagate(False)
        self.browse_entry = CTkEntry(self, corner_radius=100)
        self.browse_entry.place(y = 200, x = 950)
        #Search button
        self.search_button = tk.CTkButton(self, text="Search", width= 540, command=self.run)
        self.search_button.place(y = 350, x = 10)
        #Save result block
        self.save_file_path_label = tk.CTkLabel(self, text="Path where save result:")
        self.save_file_path_label.place(y=250, x=660)
        self.save_file_path_button = tk.CTkButton(self, text="Browse", command=self.save_result)
        self.save_file_path_button.place(y=250, x=800)
        self.save_file_path_entry = tk.CTkEntry(self, corner_radius=100)
        self.save_file_path_entry.place(y=250, x=950)
        self.estimated_completion_label = CTkLabel(self, text="")
        self.estimated_completion_label.place(y=270, x=30)
        #self.search_button.grid_propagate(False)
        self.progress_var = tk.DoubleVar()
        self.progress_bar = tk.CTkProgressBar(self, variable=self.progress_var, width=1050, height=30)
        self.progress_label = CTkLabel(self, text="")
        self.progress_label.place(y=300, x=520)
        self.progress_bar.place(y=300, x=30)
        # Quit button
        self.quit_button = tk.CTkButton(self, text="Quit", width= 540, command=self.quit)
        self.quit_button.place(y = 350, x = 555)
        #self.quit_button.grid_propagate(False)
        self.console_output = tk.CTkTextbox(self, width= 1050, height= 200, state='disabled')
        self.console_output.place(y = 400, x = 30)
        sys.stdout = ConsoleOutput(self.console_output)

    def ask_delete_debug_files(self):
        #popup = ToplevelWindow(self, self.set_delete_debug_files)
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = ToplevelWindow(self.delete_debug_files)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it


    #def set_delete_debug_files(self):
        #self.delete_debug_files = True

    def delete_debug_files(self):
        debug_folder = "mac_debug"

        if os.path.exists(debug_folder):
            try:
                logger.handlers[0].close()
                logger.removeHandler(logger.handlers[0])
            # Use the system-level command to delete the directory and its contents
                subprocess.run(["cmd", "/c", "rd", "/s", "/q", debug_folder])
                print(f"Debug folder '{debug_folder}' has been deleted.")
            except Exception as e:
                print(f"An error occurred while deleting folder '{debug_folder}': {str(e)}")
        else:
            print(f"Debug folder '{debug_folder}' not found.")



    def toggle_entry_or_browse(self):
        if self.var.get() == 1:
            self.mac_address_entry.place(y = 150, x = 950)
            self.mac_address_label.place(y = 150, x = 590)
            self.mac_address_entry_radio.place(y = 150, x = 790)
            self.mac_address_browse_radio.place(y = 150, x = 860)
            self.mac_address_browse_button.place_forget()
            self.mac_address_browse_entry.place_forget()

        else:
            self.mac_address_entry.place_forget()
            self.mac_address_label.place(y = 150, x = 440)
            self.mac_address_entry_radio.place(y = 150, x = 640)
            self.mac_address_browse_radio.place(y = 150, x= 710)
            self.mac_address_browse_button.place(y = 150, x = 790)
            self.mac_address_browse_entry.place(y = 150, x = 950)

    def browse_file_mac(self):
        # open file dialog to select file
        mac_file_path_raw = filedialog.askopenfilename()
        self.mac_file_path = mac_file_path_raw
        # set text of entry widget to selected file path
        self.mac_address_browse_entry.delete(0, tk.END)
        self.mac_address_browse_entry.insert(0, mac_file_path_raw)

    def browse_device_file(self):
        device_file_path_raw = filedialog.askopenfilename()
        self.device_file_path = device_file_path_raw
        self.browse_entry.delete(0, tk.END)
        self.browse_entry.insert(0, device_file_path_raw)

    def save_result(self):
        save_file_path_raw= filedialog.asksaveasfilename(defaultextension="")
        self.save_file_path = save_file_path_raw
        current_datetime = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        self.save_file_path_with_datetime = f"{self.save_file_path}_{current_datetime}"

        self.save_file_path_entry.delete(0, tk.END)
        self.save_file_path_entry.insert(0, self.save_file_path_with_datetime)

    def write_output_to_file(self, device, output):
        # Create the "mac_debug" folder if it doesn't exist
        if not os.path.exists("mac_debug"):
            os.mkdir("mac_debug")
    
        # Write the output to a file
        file_path_debug_output = os.path.join("mac_debug", f"{device}_output.txt")
        with open(file_path_debug_output, "w") as file:
            file.write(output)
        print(f"Port status not found in output for device {device}. Output written to file.")

    def animate_searching_dots(self):
        print("Searching", end="", flush=True)
        for _ in range(3):
            time.sleep(0.5)
            print(".", end="", flush=True)
        print("\b\b\b   \b\b\b", end="", flush=True)  # Clear dots animation
    


    def run(self):
        self.t1 = threading.Thread(target=self.search, )
        self.t1.start()

    # Create a button to start the search
    def search(self):
        self.search_button.configure(state="disabled")
        self.console_output.delete("1.0", "end")
        time.sleep(5)
        def convert_ports(port_raw):
            port = re.sub(r"(GigabitEthernet|GE|Gi|gi)", "GE", port_raw, flags=re.IGNORECASE)
            port_comware = re.sub(r"(GigabitEthernet|GE|Gi|gi)", "Gi", port_raw_comware, flags=re.IGNORECASE)
            #print(f"You entered Comware port: {port_comware}")
            return port, port_comware
        

        def convert_mac_address(mac_address):
                mac_address_st1 = mac_address.lower()
                mac_address_unsub = re.sub(r"[:-]", "", mac_address_st1)
                hex_digits = re.findall(r"[0-9a-fA-F]{2}", mac_address_unsub)
                converted_mac = "-".join([f"{hex_digits[i]}{hex_digits[i + 1]}" for i in range(0, len(hex_digits), 2)])
                return converted_mac
        while True:
            username = self.username_entry.get()
            if not username:
                print("Please provide a username")
                continue
            break
        while True:
            password = self.password_entry.get()
            if not password:
                print("Please provide a password")
                continue
            break
        # Prompt user for physical port and mac address to search for
        while True:
            port_raw = self.port_entry.get()
            port_raw_comware = port_raw
            if not port_raw:
                print("Please provide a physical port of device")
                continue
            break
        port = convert_ports(port_raw)
        port_comware = convert_ports(port_raw_comware)
        while True:
            if self.var.get() == 1:
                mac_address_raw_un = self.mac_address_entry.get()
                converted_mac_addresses = []
                try:
                    converted_mac = convert_mac_address(mac_address_raw_un)
                    converted_mac_addresses.append(converted_mac)
                    break
                except:
                    print("Failed to convert MAC address")
                    break
            else:
                try:
                    #mac_address = None
                    mac_workbook = openpyxl.load_workbook(self.mac_file_path)
                    mac_worksheet = mac_workbook.active
                except PermissionError:
                    print(f"Permission denied: Cannot save the corrected Excel file. Make sure the file is not opened by another application.")
                    return

                header_row_mac = mac_worksheet[1]
                header_names_mac = [cell.value for cell in header_row_mac]
                mac_address_col_index = header_names_mac.index("mac-address") if "mac-address" in header_names_mac else None
                if mac_address_col_index is None:
                    print("Column 'mac-address' not found in the Excel sheet.")
                    return
                # Create a list of mac address to connect to
                converted_mac_addresses = []
                for row in mac_worksheet.iter_rows(min_row=2, max_col=mac_address_col_index+1):
                    device_mac = row[mac_address_col_index].value
                    try:
                        if device_mac is not None:
                            converted_mac = convert_mac_address(device_mac)
                            converted_mac_addresses.append(converted_mac)
                    except:
                        print(f"Failed to convert MAC address: {device_mac}")
                #print("Converted MAC addresses:")
                #for converted_mac in converted_mac_addresses:
                    #print(converted_mac)
                    #mac_address=converted_mac
                    #return mac_address
                    #continue
                break
        print(f"You entered port:{port[0]}")
        device_file_path = self.device_file_path
        if not device_file_path:
            print("Please provide a path to a file with device IP addresses")
            return

        # file_path = input("Enter path to file with device IP addresses: ")
        # file_path=self.browse_file(file_path)
        try:
            device_workbook = openpyxl.load_workbook(self.device_file_path)
            device_worksheet = device_workbook.active
        except PermissionError:
            print(f"Permission denied: Cannot save the corrected Excel file. Make sure the file is not opened by another application.")
            return

        header_row_device = device_worksheet[1]
        header_names_device = [cell.value for cell in header_row_device]
        device_ip_col_index = header_names_device.index("MGMNT IP") if "MGMNT IP" in header_names_device else None
        if device_ip_col_index is None:
                    print("Column 'MGMNT IP' not found in the Excel sheet.")
                    return

        # Create a list of devices to connect to
        devices = []
        for row in device_worksheet.iter_rows(min_row=2, max_col=device_ip_col_index+1):
            device_ip = row[device_ip_col_index].value
            devices.append(device_ip)


        # Create SSH client object
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.load_system_host_keys(filename=None)
        #print(f"{mac_address}")

        # Initialize progress bar
        total_devices = len(devices)
        current_device = 0

        # Loop through devices and search for MAC address on specified port
        start_time = datetime.datetime.now()
        time_per_device = (datetime.datetime.now() - start_time) / current_device if current_device > 0 else datetime.timedelta(0)
        estimated_completion_time = start_time + time_per_device * (total_devices - current_device)
        results = []
        self.t1.join
        for device in devices:
            if device is None:
                break
            try:
                current_device += 1
                progress_percent = (current_device / total_devices) * 100
                self.per = str(int(progress_percent))
                self.progress_var.set(float(progress_percent) / 100)
                self.progress_label.configure(text=self.per + '%' )
                self.appearance_mode = CTk._get_appearance_mode(self)
                #print(f"{self.appearance_mode}")
                if self.appearance_mode == "light":
                    self.progress_label.configure(fg_color = "#989ca4", bg_color = "#989ca4")
                    if int(self.per) > 44:
                        self.progress_label.configure(fg_color="#3b8ed0", bg_color="#3b8ed0")
                    else:
                        self.progress_label.configure(fg_color="#989ca4", bg_color="#989ca4")
                else:
                    self.progress_label.configure(fg_color="#504c54", bg_color="#504c54")
                    if int(self.per) > 44:
                        self.progress_label.configure(fg_color="#1f6aa5", bg_color="#1f6aa5")
                    else:
                        self.progress_label.configure(fg_color="#504c54", bg_color="#504c54")
                self.update_idletasks()
                time_per_device = (datetime.datetime.now() - start_time) / current_device if current_device > 0 else datetime.timedelta(0)
                estimated_completion_time = start_time + time_per_device * (total_devices - current_device)
                self.estimated_completion_label.configure(text=f"Estimated Completion Time: {estimated_completion_time.strftime('%Y-%m-%d %H:%M:%S')}")
                ssh.connect(hostname=device, username=username, password=password, timeout=10)
                print(f"Connected to {device}")
                dot_animation = ["", ".", "..", "..."]
  
                searching_line = self.console_output.index("2.0")  # Get the index of the "Searching" line
                dot_count = 0  # Initialize the dot count
                while True:
                    self.console_output.delete(searching_line, searching_line + " lineend")  # Clear the "Searching" line
                    self.console_output.insert("end", "Searching" + dot_animation[dot_count % 4], "searching")
                    dot_count += 1
                    self.update_idletasks()
                    time.sleep(0.5)
                    if not self.console_output.get(searching_line, searching_line + " lineend").startswith("Searching"):
                        break
                # Open a shell
                shell = ssh.invoke_shell()
                time.sleep(2.6)
                commands = [
                    f"display mac-address | i {port[0]}",
                    f"display interface {port_comware[1]} | i Description",
                    f'display interface {port_comware[1]} | i "Current state"',
                    f"display current-configuration | i sysname"
                ]
                # Send commands to the shell
                for command in commands:
                    shell.send(command + '\n')
                    time.sleep(2.6)

                # Read the output from the shell
                output = ''
                while shell.recv_ready():
                    output += shell.recv(4096).decode('utf-8')
                    time.sleep(2.6)
                time.sleep(1)
                status_match = re.search(r'Current state:\s+(\S+)', output)
                if status_match:
                    port_status = status_match.group(1)
                    if port_status in ["UP", "DOWN", "Administratively"]:
                        if port_status in ["UP"]:
                            result_text = "Not found corresponding MAC address on device."
                            converted_macs_not_found = []
                            for converted_mac in converted_mac_addresses:
                                if converted_mac in output:
                                    self.console_output.delete(searching_line, searching_line + " lineend")  # Clear the "Searching" line
                                    self.console_output.insert("end", result_text + "\n")
                                    match = re.search(r'sysname\s+\S+\s+(\S+)', output)
                                    description = re.search(r'Description:\s+(\S+)', output)
                                    if match:
                                        sysname = match.group(1)
                                        description = description.group(1)
                                        print(
                                            f"MAC address {converted_mac} found on device {sysname} ({device}) on port {port[0]} with description:{description}"
                                            )
                                        #print(f"Port {port[0]} description: {description}")
                                    results.append((device, sysname, converted_mac, port[0], description, port_status))
                                    break  # No need to search other MAC addresses once a match is found
                                else:
                                    converted_macs_not_found.append(converted_mac)
                            if not converted_macs_not_found:
                                self.console_output.delete(searching_line, searching_line + " lineend")  # Clear the "Searching" line
                                self.console_output.insert("end", result_text + "\n")

                                match = re.search(r'sysname\s+\S+\s+(\S+)', output)
                                if match:
                                    sysname = match.group(1)
                                    description = re.search(r'Description:\s+(\S+)', output)
                                    if description:
                                        description = description.group(1)
                                self.results.append((device, sysname, None, port[0], description, port_status))

                        else:
                            print(f"Port {port[0]} is {port_status}. Skipping...")
                            match = re.search(r'sysname\s+\S+\s+(\S+)', output)
                            if match:
                                sysname = match.group(1)
                                #print(f"{sysname}")
                                description = re.search(r'Description:\s+(\S+)', output)
                            if description:
                                description = description.group(1)
                            results.append((device, sysname, None, port[0], description, port_status))
                    else:
                        print(f"Unknown port status '{port_status}'. Skipping...")
                else:
                    self.write_output_to_file(device, output)
                #print(f"{output}")

            except Exception as e:
                print(f"An error occurred while processing device {device}: {str(e)}")
                    # print("\n")
                self.write_output_to_file(device, output)
                time.sleep(4)
            except paramiko.AuthenticationException:
                print(f"Authentication failed for {device}")
            except paramiko.SSHException:
                print(f"Unable to establish SSH connection to {device}")
            except paramiko.ssh_exception.NoValidConnectionsError:
                print(f"Unable to connect to {device}")
            except Exception as e:
                print(e)
            finally:
                ssh.close()
        self.progress_var.set(100.0)
        self.update_idletasks()
        end_time = datetime.datetime.now()
        elapsed_time = end_time - start_time
        elapsed_time_seconds = elapsed_time.total_seconds()
        print(f"Search completed in {elapsed_time_seconds:.2f} seconds")
        # Create an Excel workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        self.ask_delete_debug_files()
        #if self.delete_debug_files:
            #self.delete_debug_files()

        # Write headers to the worksheet
        ws.append(["MGMNT IP", "Sysname", "MAC Address", "Interface", "Description", "Port Status"])

        # Write data to the worksheet
        for result in results:
            ws.append([str(value) for value in result])

        try:
            wb.save(f"{self.save_file_path_with_datetime}.xlsx")
            print(f"Results saved to {self.save_file_path_with_datetime}.xlsx")
        except Exception as e:
            print(f"An error occurred while saving the Excel file: {str(e)}")
        self.search_button.configure(state="normal")
        # Console output
        # self.console = Console(self)
        #threads = []
        #for device in devices:
            #thread = threading.Thread(target=connect_ssh, args=(device,))
            #thread.start()
            #threads.append(thread)

        # Wait for all threads to complete
        #for thread in threads:
            #thread.join()
    def quit(self):
        self.destroy()


if __name__ == "__main__":
    app = Application()
    app.mainloop()
